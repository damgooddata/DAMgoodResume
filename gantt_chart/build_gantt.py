"""
Build a dynamic, visually polished Gantt chart in Excel from the sample project dataset.
"""

import csv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
INPUT_CSV = "/home/user/DAMgoodResume/gantt_chart/sample_project_data.csv"
OUTPUT_XLSX = "/home/user/DAMgoodResume/gantt_chart/DAMgood_Gantt_Chart.xlsx"

# Phase colors (fill for bars and phase labels)
PHASE_COLORS = {
    "Planning":   "1B4F72",
    "Discovery":  "2E86C1",
    "Design":     "2ECC71",
    "Build":      "E67E22",
    "Testing":    "E74C3C",
    "Deployment": "8E44AD",
    "Closeout":   "17A589",
}

PHASE_BAR_FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in PHASE_COLORS.items()}

# Milestone diamond color
MILESTONE_FILL = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
MILESTONE_FONT = Font(bold=True, color="000000", size=11)

# Status colors for the status column
STATUS_FILLS = {
    "Complete":     PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "In Progress":  PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
    "Not Started":  PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid"),
}

STATUS_FONTS = {
    "Complete":     Font(color="FFFFFF", bold=True, size=10),
    "In Progress":  Font(color="FFFFFF", bold=True, size=10),
    "Not Started":  Font(color="2C3E50", bold=False, size=10),
}

# Priority colors
PRIORITY_FILLS = {
    "High":   PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "Medium": PatternFill(start_color="F5B041", end_color="F5B041", fill_type="solid"),
}
PRIORITY_FONTS = {
    "High":   Font(color="FFFFFF", bold=True, size=10),
    "Medium": Font(color="000000", bold=False, size=10),
}

# General styles
HEADER_FILL = PatternFill(start_color="1C2833", end_color="1C2833", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
PHASE_ROW_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
PHASE_ROW_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
DATA_FONT = Font(size=10, name="Calibri", color="2C3E50")
DATA_FONT_BOLD = Font(size=10, name="Calibri", color="2C3E50", bold=True)
LIGHT_FILL_A = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
LIGHT_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
TODAY_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

PROGRESS_FILL = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")

# Gantt bar opacity variants (lighter shade for incomplete portion)
def lighter_hex(hex_color, factor=0.45):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
def load_tasks(path):
    tasks = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row["task_id"] = int(row["task_id"])
            row["start_date"] = datetime.strptime(row["start_date"], "%Y-%m-%d").date()
            row["end_date"] = datetime.strptime(row["end_date"], "%Y-%m-%d").date()
            row["duration_days"] = int(row["duration_days"])
            row["pct_complete"] = int(row["pct_complete"]) if row["pct_complete"] else 0
            row["is_milestone"] = row["is_milestone"].strip().lower() == "yes"
            tasks.append(row)
    return tasks


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
def build_gantt(tasks):
    wb = Workbook()

    # -----------------------------------------------------------------------
    # SHEET 1: Task Data (reference table)
    # -----------------------------------------------------------------------
    ws_data = wb.active
    ws_data.title = "Task Data"
    ws_data.sheet_properties.tabColor = "1B4F72"

    data_headers = [
        "ID", "Phase", "Task", "Assignee", "Role",
        "Start", "End", "Days", "Predecessor", "Dep Type",
        "% Complete", "Status", "Priority", "Milestone", "Notes"
    ]
    data_widths = [5, 14, 42, 18, 20, 12, 12, 6, 14, 9, 11, 13, 9, 10, 50]

    for ci, (header, width) in enumerate(zip(data_headers, data_widths), 1):
        cell = ws_data.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws_data.column_dimensions[get_column_letter(ci)].width = width

    ws_data.row_dimensions[1].height = 30

    for ri, t in enumerate(tasks, 2):
        vals = [
            t["task_id"], t["phase"], t["task_name"], t["assignee"], t["role"],
            t["start_date"], t["end_date"], t["duration_days"],
            t["predecessor"], t["dependency_type"],
            t["pct_complete"] / 100, t["status"], t["priority"],
            "Yes" if t["is_milestone"] else "", t["notes"]
        ]
        row_fill = LIGHT_FILL_A if ri % 2 == 0 else LIGHT_FILL_B
        for ci, val in enumerate(vals, 1):
            cell = ws_data.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 15))

            if ci == 6 or ci == 7:  # dates
                cell.number_format = "MMM DD, YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 11:  # pct
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (1, 8, 10, 14):
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Status coloring
            if ci == 12:
                sf = STATUS_FILLS.get(val)
                if sf:
                    cell.fill = sf
                    cell.font = STATUS_FONTS.get(val, DATA_FONT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Priority coloring
            if ci == 13:
                pf = PRIORITY_FILLS.get(val)
                if pf:
                    cell.fill = pf
                    cell.font = PRIORITY_FONTS.get(val, DATA_FONT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_data.row_dimensions[ri].height = 22

    # Freeze top row
    ws_data.freeze_panes = "A2"
    # Auto-filter
    ws_data.auto_filter.ref = f"A1:O{len(tasks)+1}"

    # -----------------------------------------------------------------------
    # SHEET 2: Gantt Chart
    # -----------------------------------------------------------------------
    ws = wb.create_sheet("Gantt Chart")
    ws.sheet_properties.tabColor = "E67E22"

    # Determine date range for chart columns
    all_starts = [t["start_date"] for t in tasks]
    all_ends = [t["end_date"] for t in tasks]
    chart_start = min(all_starts) - timedelta(days=min(all_starts).weekday())  # start on Monday
    chart_end = max(all_ends) + timedelta(days=(6 - max(all_ends).weekday()))  # end on Sunday
    # Add a small buffer
    chart_end += timedelta(days=7)
    total_days = (chart_end - chart_start).days + 1

    # Left-side columns: ID, Phase, Task, Assignee, Status, % Complete
    left_headers = ["ID", "Phase", "Task", "Assignee", "Status", "Progress"]
    left_widths = [5, 13, 38, 16, 12, 10]
    gantt_col_start = len(left_headers) + 1  # first date column

    # ----- ROW 1: Month headers -----
    # ----- ROW 2: Week headers -----
    # ----- ROW 3: Column headers for left side + day-level date ticks -----

    # Row 1 & 2: empty for left columns
    for ci in range(1, gantt_col_start):
        for ri in (1, 2, 3):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

    # Left-side header labels in row 3
    for ci, (h, w) in enumerate(zip(left_headers, left_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Date columns: one column per day
    # We'll use weekly granularity to keep it manageable
    # Actually, let's use weekly columns for a cleaner chart
    # Each column = 1 week (Mon-Sun)
    weeks = []
    d = chart_start
    while d <= chart_end:
        week_end = d + timedelta(days=6)
        weeks.append((d, min(week_end, chart_end)))
        d = week_end + timedelta(days=1)

    # Row 1: Month headers (merged across weeks in that month)
    month_spans = []
    current_month = None
    span_start = None
    for wi, (ws_date, we_date) in enumerate(weeks):
        month_key = (ws_date.year, ws_date.month)
        if month_key != current_month:
            if current_month is not None:
                month_spans.append((current_month, span_start, wi - 1))
            current_month = month_key
            span_start = wi
        if wi == len(weeks) - 1:
            month_spans.append((current_month, span_start, wi))

    MONTH_COLORS = ["1B4F72", "2E86C1", "1B7A4F", "E67E22", "8E44AD", "C0392B", "17A589"]
    for mi, (month_key, ms, me) in enumerate(month_spans):
        col_s = gantt_col_start + ms
        col_e = gantt_col_start + me
        month_name = datetime(month_key[0], month_key[1], 1).strftime("%B %Y")
        cell = ws.cell(row=1, column=col_s, value=month_name)
        mc = MONTH_COLORS[mi % len(MONTH_COLORS)]
        cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        cell.fill = PatternFill(start_color=mc, end_color=mc, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if col_e > col_s:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        # Fill merged area
        for c in range(col_s + 1, col_e + 1):
            fc = ws.cell(row=1, column=c)
            fc.fill = PatternFill(start_color=mc, end_color=mc, fill_type="solid")
            fc.border = THIN_BORDER

    # Row 2: Week start dates
    for wi, (ws_date, we_date) in enumerate(weeks):
        col = gantt_col_start + wi
        cell = ws.cell(row=2, column=col, value=ws_date.strftime("%b %d"))
        cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 7

    # Row 3: just fill the date header area
    for wi in range(len(weeks)):
        col = gantt_col_start + wi
        cell = ws.cell(row=3, column=col, value="")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Set row heights for headers
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 24

    # -----------------------------------------------------------------------
    # Populate task rows
    # -----------------------------------------------------------------------
    data_row = 4  # first data row
    current_phase = None

    for t in tasks:
        # Insert phase separator row
        if t["phase"] != current_phase:
            current_phase = t["phase"]
            phase_color = PHASE_COLORS.get(current_phase, "2C3E50")
            phase_fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type="solid")

            # Merge left columns for phase name
            cell = ws.cell(row=data_row, column=1, value="")
            cell.fill = phase_fill
            cell.border = THIN_BORDER
            cell = ws.cell(row=data_row, column=2, value=current_phase.upper())
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = phase_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER
            for ci in range(3, gantt_col_start + len(weeks)):
                c = ws.cell(row=data_row, column=ci)
                c.fill = phase_fill
                c.border = THIN_BORDER
            ws.row_dimensions[data_row].height = 26
            data_row += 1

        # Task row
        row_fill = LIGHT_FILL_A if data_row % 2 == 0 else LIGHT_FILL_B

        # Left columns
        left_vals = [
            t["task_id"],
            "",  # phase (shown in separator)
            ("  \u25C6  " + t["task_name"]) if t["is_milestone"] else ("     " + t["task_name"]),
            t["assignee"],
            t["status"],
            t["pct_complete"] / 100,
        ]

        for ci, val in enumerate(left_vals, 1):
            cell = ws.cell(row=data_row, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.fill = row_fill

            if ci == 1:
                cell.font = Font(size=9, color="95A5A6", name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 3:  # task name
                if t["is_milestone"]:
                    cell.font = Font(size=10, bold=True, color="D4AC0D", name="Calibri")
                else:
                    cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="center")
            elif ci == 4:
                cell.font = Font(size=9, color="5D6D7E", name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:  # status
                sf = STATUS_FILLS.get(val)
                if sf:
                    cell.fill = sf
                    cell.font = STATUS_FONTS.get(val, DATA_FONT)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 6:  # progress
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = DATA_FONT_BOLD
            else:
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="center")

        # Gantt bar columns
        phase_hex = PHASE_COLORS.get(t["phase"], "2C3E50")
        bar_fill = PatternFill(start_color=phase_hex, end_color=phase_hex, fill_type="solid")
        light_hex = lighter_hex(phase_hex, 0.5)
        bar_fill_light = PatternFill(start_color=light_hex, end_color=light_hex, fill_type="solid")
        progress_done_fill = PatternFill(start_color=phase_hex, end_color=phase_hex, fill_type="solid")

        task_start = t["start_date"]
        task_end = t["end_date"]
        pct = t["pct_complete"]
        task_duration = (task_end - task_start).days + 1
        completed_days = int(task_duration * pct / 100)

        for wi, (week_start, week_end) in enumerate(weeks):
            col = gantt_col_start + wi
            cell = ws.cell(row=data_row, column=col)
            cell.border = Border(
                left=Side(style="thin", color="EAECEE"),
                right=Side(style="thin", color="EAECEE"),
                top=Side(style="thin", color="D5D8DC"),
                bottom=Side(style="thin", color="D5D8DC"),
            )
            cell.fill = row_fill

            # Check if this week overlaps with the task
            overlap_start = max(task_start, week_start)
            overlap_end = min(task_end, week_end)

            if overlap_start <= overlap_end:
                if t["is_milestone"]:
                    cell.value = "\u25C6"
                    cell.font = Font(size=14, bold=True, color="D4AC0D", name="Calibri")
                    cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Determine if this week portion is in completed or remaining
                    days_from_start = (overlap_start - task_start).days
                    days_from_start_end = (overlap_end - task_start).days

                    if days_from_start_end < completed_days:
                        # Fully in completed portion
                        cell.fill = progress_done_fill
                        cell.value = ""
                    elif days_from_start >= completed_days:
                        # Fully in remaining portion
                        cell.fill = bar_fill_light
                        cell.value = ""
                    else:
                        # Mixed - show as completed (darker)
                        cell.fill = progress_done_fill
                        cell.value = ""

                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[data_row].height = 24
        data_row += 1

    # -----------------------------------------------------------------------
    # Add a legend below the chart
    # -----------------------------------------------------------------------
    legend_row = data_row + 2
    ws.cell(row=legend_row, column=2, value="LEGEND").font = Font(bold=True, size=12, color="2C3E50", name="Calibri")
    legend_row += 1

    # Phase colors
    for phase, color in PHASE_COLORS.items():
        cell = ws.cell(row=legend_row, column=2)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = THIN_BORDER
        cell = ws.cell(row=legend_row, column=3, value=phase)
        cell.font = Font(size=10, color="2C3E50", name="Calibri")
        cell.alignment = Alignment(vertical="center")

        # Show lighter shade
        lc = lighter_hex(color, 0.5)
        cell2 = ws.cell(row=legend_row, column=4)
        cell2.fill = PatternFill(start_color=lc, end_color=lc, fill_type="solid")
        cell2.border = THIN_BORDER
        cell3 = ws.cell(row=legend_row, column=5, value="Remaining")
        cell3.font = Font(size=9, color="7F8C8D", name="Calibri", italic=True)
        legend_row += 1

    legend_row += 1
    # Status legend
    for status, fill in STATUS_FILLS.items():
        cell = ws.cell(row=legend_row, column=2)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell = ws.cell(row=legend_row, column=3, value=status)
        cell.font = Font(size=10, color="2C3E50", name="Calibri")
        legend_row += 1

    legend_row += 1
    # Milestone legend
    cell = ws.cell(row=legend_row, column=2, value="\u25C6")
    cell.font = Font(size=14, bold=True, color="D4AC0D", name="Calibri")
    cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=legend_row, column=3, value="Milestone")
    cell.font = Font(size=10, color="2C3E50", name="Calibri")

    # Freeze panes: freeze left columns and top 3 header rows
    ws.freeze_panes = ws.cell(row=4, column=gantt_col_start)

    # -----------------------------------------------------------------------
    # SHEET 3: Summary Dashboard
    # -----------------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "27AE60"

    ws_sum.column_dimensions["A"].width = 3
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 14
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 14
    ws_sum.column_dimensions["F"].width = 14
    ws_sum.column_dimensions["G"].width = 14

    # Title
    ws_sum.merge_cells("B2:G2")
    cell = ws_sum.cell(row=2, column=2, value="PROJECT SUMMARY DASHBOARD")
    cell.font = Font(bold=True, size=16, color="1C2833", name="Calibri")
    cell.alignment = Alignment(horizontal="center")

    # Overall stats
    total = len(tasks)
    complete = sum(1 for t in tasks if t["status"] == "Complete")
    in_progress = sum(1 for t in tasks if t["status"] == "In Progress")
    not_started = sum(1 for t in tasks if t["status"] == "Not Started")
    milestones = sum(1 for t in tasks if t["is_milestone"])
    overall_pct = sum(t["pct_complete"] for t in tasks) / total if total else 0

    stats = [
        ("Total Tasks", total),
        ("Completed", complete),
        ("In Progress", in_progress),
        ("Not Started", not_started),
        ("Milestones", milestones),
        ("Overall Progress", overall_pct / 100),
    ]

    row = 4
    stat_header_fill = PatternFill(start_color="1C2833", end_color="1C2833", fill_type="solid")
    for ci, h in enumerate(["Metric", "Value"], 2):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = stat_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1
    for label, val in stats:
        cell = ws_sum.cell(row=row, column=2, value=label)
        cell.font = DATA_FONT_BOLD
        cell.border = THIN_BORDER
        cell.fill = LIGHT_FILL_A if row % 2 == 0 else LIGHT_FILL_B
        cell = ws_sum.cell(row=row, column=3, value=val)
        cell.font = Font(size=12, bold=True, color="1B4F72", name="Calibri")
        cell.border = THIN_BORDER
        cell.fill = LIGHT_FILL_A if row % 2 == 0 else LIGHT_FILL_B
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if label == "Overall Progress":
            cell.number_format = "0%"
        row += 1

    # Phase breakdown
    row += 1
    ws_sum.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    cell = ws_sum.cell(row=row, column=2, value="PHASE BREAKDOWN")
    cell.font = Font(bold=True, size=13, color="1C2833", name="Calibri")
    cell.alignment = Alignment(horizontal="center")
    row += 1

    phase_headers = ["Phase", "Tasks", "Complete", "In Progress", "Not Started", "Avg Progress"]
    for ci, h in enumerate(phase_headers, 2):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = stat_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    seen_phases = []
    for t in tasks:
        if t["phase"] not in seen_phases:
            seen_phases.append(t["phase"])

    for phase in seen_phases:
        phase_tasks = [t for t in tasks if t["phase"] == phase]
        pt = len(phase_tasks)
        pc = sum(1 for t in phase_tasks if t["status"] == "Complete")
        pip = sum(1 for t in phase_tasks if t["status"] == "In Progress")
        pns = sum(1 for t in phase_tasks if t["status"] == "Not Started")
        pavg = sum(t["pct_complete"] for t in phase_tasks) / pt if pt else 0

        vals = [phase, pt, pc, pip, pns, pavg / 100]
        rf = LIGHT_FILL_A if row % 2 == 0 else LIGHT_FILL_B
        phase_color = PHASE_COLORS.get(phase, "2C3E50")
        for ci, val in enumerate(vals, 2):
            cell = ws_sum.cell(row=row, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 2:
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.fill = rf
                cell.font = DATA_FONT
            if ci == 7:
                cell.number_format = "0%"
                cell.font = DATA_FONT_BOLD
        row += 1

    # Team workload
    row += 1
    ws_sum.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws_sum.cell(row=row, column=2, value="TEAM WORKLOAD")
    cell.font = Font(bold=True, size=13, color="1C2833", name="Calibri")
    cell.alignment = Alignment(horizontal="center")
    row += 1

    team_headers = ["Assignee", "Tasks", "Total Days", "Avg Progress"]
    for ci, h in enumerate(team_headers, 2):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = stat_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    assignees = {}
    for t in tasks:
        a = t["assignee"]
        if a not in assignees:
            assignees[a] = {"count": 0, "days": 0, "pct_sum": 0}
        assignees[a]["count"] += 1
        assignees[a]["days"] += t["duration_days"]
        assignees[a]["pct_sum"] += t["pct_complete"]

    for name, info in sorted(assignees.items()):
        rf = LIGHT_FILL_A if row % 2 == 0 else LIGHT_FILL_B
        vals = [name, info["count"], info["days"], (info["pct_sum"] / info["count"]) / 100]
        for ci, val in enumerate(vals, 2):
            cell = ws_sum.cell(row=row, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.fill = rf
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 2:
                cell.font = DATA_FONT_BOLD
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 5:
                cell.number_format = "0%"
                cell.font = DATA_FONT_BOLD
            else:
                cell.font = DATA_FONT
        row += 1

    ws_sum.freeze_panes = "A3"

    # -----------------------------------------------------------------------
    # Set Gantt Chart as the active/default sheet
    # -----------------------------------------------------------------------
    wb.active = wb.sheetnames.index("Gantt Chart")

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"Gantt chart saved to: {OUTPUT_XLSX}")
    print(f"  - {len(tasks)} tasks across {len(seen_phases)} phases")
    print(f"  - {len(weeks)} weekly columns spanning {chart_start} to {chart_end}")
    print(f"  - 3 sheets: Task Data, Gantt Chart, Summary")


if __name__ == "__main__":
    tasks = load_tasks(INPUT_CSV)
    build_gantt(tasks)
